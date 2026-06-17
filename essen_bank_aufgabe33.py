"""
Essen Bank – Aufgabe 3.3: Output Floor gem. CRR III
====================================================
Berechnung von IRBA-RWA und KSA-RWA für 50 Unternehmenskredite
sowie Szenarioanalyse (a/b/c) zum Output Floor.

Rechtsgrundlagen:
  - Art. 153 CRR  : IRBA-Risikogewicht (Vasicek-Formel)
  - Art. 160 CRR  : PD-Floor 0,03 %
  - Art. 161 CRR  : LGD
  - Art. 162 CRR  : Restlaufzeit M
  - Art. 501 CRR  : KMU-Korrektur der Korrelation
  - Art. 122 CRR III : KSA-Risikogewichte
  - Art. 465 CRR III : Output Floor (72,5 % ab Endstufe 2033)

Abhängigkeiten: pip install scipy openpyxl
"""

import math
from scipy.stats import norm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════════════════════════════════════════
# 1. PORTFOLIODATEN
# ═══════════════════════════════════════════════════════════════════════════════
# Felder: (ID, Name, Branche, PD_%, LGD_%, M_Jahre, EAD_Mio, KMU, S_Mio, Rating)
# KMU  : True/False – KMU-Korrektur gem. Art. 501 CRR
# S    : Jahresumsatz in Mio. € (nur relevant bei KMU=True)
# Rating: "A" / "BBB" / "BB" / "B" / "unrated"

PORTFOLIO = [
    (1,  "Metallverarbeitung Ruhr AG",          "Metall",        0.42, 38, 2.5, 22,  False, None, "BBB"),
    (2,  "Hotelkette Ruhr GmbH",                "Hotel",         0.85, 35, 4.0, 20,  False, None, "BB"),
    (3,  "Logistik Ruhr GmbH",                  "Logistik",      0.12, 36, 2.5, 12,  False, None, "A"),
    (4,  "Wohnungsbau Ruhr AG",                 "Gewerbeimmo",   0.20, 28, 4.0, 38,  False, None, "BBB"),
    (5,  "Spedition Ruhr GmbH",                 "Logistik",      0.65, 40, 2.5, 11,  False, None, "BB"),
    (6,  "Pflegeheim Ruhr GmbH",                "Gesundheit",    0.33, 30, 4.0, 14,  True,  30,   "BBB"),
    (7,  "Bauträger Ruhr GmbH",                 "Bau",           1.90, 35, 4.0, 22,  False, None, "B"),
    (8,  "Stahlhandel Essen GmbH",              "Handel",        0.18, 35, 2.5, 14,  False, None, "BBB"),
    (9,  "Fensterbau Ruhr GmbH",                "Bau",           2.60, 40, 2.5,  5,  True,  12,   "unrated"),
    (10, "Industrietechnik Ruhr AG",            "Maschinenbau",  0.25, 38, 2.5, 38,  False, None, "BBB"),
    (11, "Gastronomiebetrieb Ruhr GmbH",        "Gastronomie",   0.55, 40, 2.0,  5,  True,   8,   "BB"),
    (12, "Gewerbeimmobilien Essen KG",          "Gewerbeimmo",   0.22, 30, 4.0, 25,  False, None, "BBB"),
    (13, "Druckguss Ruhr GmbH",                 "Metall",        1.80, 42, 2.5,  9,  True,  22,   "B"),
    (14, "IT-Services Ruhr GmbH",               "IT",            0.32, 45, 2.5,  7,  True,  15,   "BBB"),
    (15, "Kunststoffwerk Ruhr GmbH",            "Industrie",     0.95, 40, 2.5, 13,  True,  28,   "BB"),
    (16, "Medizintechnik Ruhr GmbH",            "Gesundheit",    0.15, 35, 2.5, 18,  False, None, "A"),
    (17, "Einzelhandel Center Ruhr GmbH",       "Handel",        2.20, 40, 2.0,  8,  True,  18,   "B"),
    (18, "Maschinenbau Ruhr GmbH",              "Maschinenbau",  0.30, 40, 2.5, 12,  True,  32,   "BBB"),
    (19, "Reinigungsservice Ruhr GmbH",         "Dienstleistung",2.10, 38, 2.5,  4,  True,   8,   "unrated"),
    (20, "Projektentwickler Ruhr GmbH",         "Gewerbeimmo",   0.45, 36, 4.0, 22,  False, None, "BBB"),
    (21, "Textilfertigung Ruhr GmbH",           "Textil",        0.62, 45, 2.5,  9,  True,  16,   "BB"),
    (22, "Energietechnik Ruhr GmbH",            "Industrie",     0.38, 37, 2.5, 10,  True,  25,   "BBB"),
    (23, "Gastro Holding Ruhr GmbH",            "Gastronomie",   2.80, 50, 2.0,  7,  True,  14,   "B"),
    (24, "Großhandel Ruhr GmbH",                "Handel",        0.28, 34, 2.5, 11,  False, None, "BBB"),
    (25, "Wohnprojekt Ruhr KG",                 "Gewerbeimmo",   1.25, 38, 4.0, 28,  False, None, "BB"),
    (26, "Elektroinstallation Ruhr GmbH",       "Handwerk",      0.35, 35, 2.5,  5,  True,  10,   "BBB"),
    (27, "Autohaus Ruhr GmbH",                  "Kfz",           1.00, 40, 2.5, 16,  False, None, "BB"),
    (28, "Lebensmittelhandel Ruhr GmbH",        "Lebensmittel",  0.35, 32, 2.5,  9,  True,  22,   "BBB"),
    (29, "Facility Management Ruhr GmbH",       "Dienstleistung",2.80, 45, 2.5,  5,  True,  10,   "unrated"),
    (30, "Bauunternehmen Ruhr GmbH",            "Bau",           0.38, 35, 2.5, 20,  False, None, "BBB"),
    (31, "Elektronik-Zulieferer Ruhr GmbH",     "Elektronik",    0.80, 40, 2.5,  6,  True,  14,   "BB"),
    (32, "Hotelbetrieb Ruhr GmbH",              "Hotel",         0.40, 40, 4.0, 16,  True,  28,   "BBB"),
    (33, "Stahlbau Ruhr GmbH",                  "Bau",           0.90, 42, 2.5,  8,  True,  19,   "BB"),
    (34, "Medienverlag Ruhr GmbH",              "Medien",        3.50, 48, 2.0,  6,  True,  16,   "B"),
    (35, "Kfz-Betrieb Ruhr GmbH",              "Kfz",           0.28, 38, 2.5,  6,  True,  18,   "BBB"),
    (36, "Büroausstatter Ruhr GmbH",            "Handel",        3.20, 42, 2.5,  5,  True,  18,   "unrated"),
    (37, "Spezialchemie Ruhr GmbH",             "Chemie",        0.48, 40, 2.5, 15,  False, None, "BBB"),
    (38, "Einzelhandel Ruhr GmbH",              "Handel",        1.20, 45, 2.5,  6,  True,  12,   "BB"),
    (39, "Logistik Essen GmbH",                 "Logistik",      2.40, 42, 2.0,  4,  True,  10,   "unrated"),
    (40, "Großbäckerei Ruhr GmbH",              "Lebensmittel",  0.42, 32, 2.5,  8,  True,  24,   "BBB"),
    (41, "Kfz-Teile Ruhr GmbH",                "Handel",        0.75, 40, 2.5,  9,  False, None, "BB"),
    (42, "Druckerei Ruhr GmbH",                 "Druck",         0.45, 40, 2.5,  6,  True,  12,   "BBB"),
    (43, "Immobilienprojekt Ruhr KG",           "Gewerbeimmo",   1.40, 38, 4.0, 24,  False, None, "BB"),
    (44, "Elektronik-Zulieferer Essen GmbH",    "Elektronik",    2.00, 45, 2.5,  5,  True,  11,   "B"),
    (45, "Zahnarztpraxis Ruhr GmbH",            "Gesundheit",    0.80, 30, 2.5,  3.5,True,   6,   "BB"),
    (46, "Modehandel Ruhr GmbH",                "Handel",        3.80, 48, 2.0,  5,  True,  13,   "unrated"),
    (47, "Stahlservice Center Ruhr GmbH",       "Metall",        3.10, 45, 2.5,  5,  True,  10,   "B"),
    (48, "Gastrobetrieb Essen GmbH",            "Gastronomie",   3.50, 45, 1.5,  3,  True,   7,   "unrated"),
    (49, "Maschinenbau Essen GmbH",             "Maschinenbau",  0.58, 38, 2.5,  8,  True,  20,   "BB"),
    (50, "Regionaler Netzbetreiber Ruhr GmbH",  "Energie",       0.10, 30, 3.0, 30,  False, None, "A"),
]

# KSA-Risikogewichte gem. Art. 122 CRR III (in Dezimal)
KSA_GEWICHTE = {
    "A":       0.50,
    "BBB":     0.75,
    "BB":      1.00,
    "B":       1.50,
    "unrated": 1.00,
}

# Output-Floor-Satz gem. Art. 465 CRR III (Endstufe ab 2033)
OUTPUT_FLOOR = 0.725


# ═══════════════════════════════════════════════════════════════════════════════
# 2. BERECHNUNGSFUNKTIONEN
# ═══════════════════════════════════════════════════════════════════════════════

def berechne_irba_rw(pd_pct, lgd_pct, M, kmu, S_mio):
    """
    Berechnet das IRBA-Risikogewicht gem. Art. 153 Abs. 1 CRR.

    Parameter
    ---------
    pd_pct  : PD in Prozent (z.B. 0.25 für 0,25 %)
    lgd_pct : LGD in Prozent (z.B. 45 für 45 %)
    M       : Restlaufzeit in Jahren
    kmu     : True wenn KMU-Korrektur anwendbar (Art. 501 CRR)
    S_mio   : Jahresumsatz in Mio. € (nur bei kmu=True relevant)

    Rückgabe
    --------
    rw      : Risikogewicht als Dezimalzahl (z.B. 0.5423 für 54,23 %)
    details : dict mit allen Zwischenergebnissen
    """

    # ── Schritt 1: PD in Dezimal + aufsichtlicher Mindest-PD ─────────────────
    # Art. 160 Abs. 1 CRR: PD >= 0,03 %
    pd  = max(pd_pct / 100, 0.0003)
    lgd = lgd_pct / 100

    # ── Schritt 2: Asset-Korrelation R ───────────────────────────────────────
    # Art. 153 Abs. 1 CRR – Formel:
    # R = 0,12 × (1 − e^(−50×PD)) / (1 − e^(−50))
    #   + 0,24 × (1 − (1 − e^(−50×PD)) / (1 − e^(−50)))
    # → R bewegt sich zwischen 12 % (sehr hohe PD) und 24 % (sehr niedrige PD)

    e1 = math.exp(-50 * pd)
    w  = (1 - e1) / (1 - math.exp(-50))   # Gewicht, abhängig von PD
    R  = 0.12 * w + 0.24 * (1 - w)

    # ── Schritt 3: KMU-Korrektur der Korrelation ─────────────────────────────
    # Art. 153 Abs. 4 CRR / Art. 501 CRR:
    # ΔR = −0,04 × (1 − (S − 5) / 45)
    # S wird auf das Intervall [5; 50] Mio. € geclampt.
    delta_R = 0.0
    S_eff   = None
    if kmu and S_mio is not None:
        S_eff   = max(min(S_mio, 50), 5)   # clamping auf [5;50]
        delta_R = -0.04 * (1 - (S_eff - 5) / 45)
    R_final = R + delta_R

    # ── Schritt 4: Laufzeitanpassung (Maturity Adjustment) ───────────────────
    # Art. 153 Abs. 1 CRR:
    # b  = (0,11852 − 0,05478 × ln(PD))²
    # MA = (1 + (M − 2,5) × b) / (1 − 1,5 × b)
    b  = (0.11852 - 0.05478 * math.log(pd)) ** 2
    MA = (1 + (M - 2.5) * b) / (1 - 1.5 * b)

    # ── Schritt 5: Vasicek-Formel – Kapitalanforderung K ─────────────────────
    # z = (Φ⁻¹(PD) + √R × Φ⁻¹(0,999)) / √(1−R)
    # K = (LGD × Φ(z) − LGD × PD) × MA
    z     = (norm.ppf(pd) + math.sqrt(R_final) * norm.ppf(0.999)) / math.sqrt(1 - R_final)
    K_roh = lgd * norm.cdf(z) - lgd * pd   # unerwarteter Verlust
    K     = K_roh * MA                      # laufzeitadjustiert

    # ── Schritt 6: Risikogewicht ──────────────────────────────────────────────
    # RW = K × 12,5  (entspricht K / 8 %)
    rw = K * 12.5

    details = {
        "pd_eff": pd, "lgd": lgd, "M": M,
        "R_base": R, "delta_R": delta_R, "R_final": R_final,
        "S_eff": S_eff, "b": b, "MA": MA,
        "z": z, "K_roh": K_roh, "K": K,
    }
    return rw, details


def berechne_ksa_rw(rating):
    """KSA-Risikogewicht gem. Art. 122 CRR III."""
    return KSA_GEWICHTE.get(rating, 1.00)


def berechne_position(eintrag):
    """
    Berechnet IRBA-RWA und KSA-RWA für eine Portfolioposition.
    Gibt ein dict mit allen Ergebnissen zurück.
    """
    id_, name, branche, pd_pct, lgd_pct, M, ead, kmu, S_mio, rating = eintrag

    # IRBA
    irba_rw, details = berechne_irba_rw(pd_pct, lgd_pct, M, kmu, S_mio)
    irba_rwa = irba_rw * ead

    # KSA
    ksa_rw  = berechne_ksa_rw(rating)
    ksa_rwa = ksa_rw * ead

    return {
        "id": id_, "name": name, "branche": branche,
        "pd_pct": pd_pct, "lgd_pct": lgd_pct, "M": M,
        "ead": ead, "kmu": kmu, "S_mio": S_mio, "rating": rating,
        # Ergebnisse
        "irba_rw":  irba_rw,
        "irba_rwa": irba_rwa,
        "ksa_rw":   ksa_rw,
        "ksa_rwa":  ksa_rwa,
        "differenz": ksa_rwa - irba_rwa,
        # Zwischenschritte
        **{f"_{k}": v for k, v in details.items()},
    }


def floor_analyse(positionen):
    """
    Berechnet Portfolio-Summen und Output-Floor-Ergebnis
    für eine gegebene Positionsliste.
    """
    irba_sum = sum(p["irba_rwa"] for p in positionen)
    ksa_sum  = sum(p["ksa_rwa"]  for p in positionen)
    ead_sum  = sum(p["ead"]      for p in positionen)
    floor_basis = OUTPUT_FLOOR * ksa_sum
    floor_rwa   = max(irba_sum, floor_basis)
    return {
        "ead_sum":     ead_sum,
        "irba_sum":    irba_sum,
        "ksa_sum":     ksa_sum,
        "floor_basis": floor_basis,
        "floor_rwa":   floor_rwa,
        "bindet":      floor_basis > irba_sum,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 3. PORTFOLIO BERECHNEN
# ═══════════════════════════════════════════════════════════════════════════════

ergebnisse = [berechne_position(p) for p in PORTFOLIO]
baseline   = floor_analyse(ergebnisse)


# ═══════════════════════════════════════════════════════════════════════════════
# 4. SZENARIEN (a / b / c)
# ═══════════════════════════════════════════════════════════════════════════════

def szenario(ergebnisse, sortkey, label):
    """
    Verkauft die 10 Positionen mit dem höchsten Wert laut sortkey.
    Gibt Ergebnis-dict zurück.
    """
    sortiert   = sorted(ergebnisse, key=lambda p: p[sortkey], reverse=True)
    verkauft   = sortiert[:10]
    verbleibend = sortiert[10:]
    analyse    = floor_analyse(verbleibend)
    return {
        "label":       label,
        "sortkey":     sortkey,
        "verkauft":    verkauft,
        "verbleibend": verbleibend,
        **analyse,
    }

sz_a = szenario(ergebnisse, "ead",       "(a) 10 größte EADs")
sz_b = szenario(ergebnisse, "irba_rwa",  "(b) 10 höchste IRBA-RWA")
sz_c = szenario(ergebnisse, "differenz", "(c) 10 größte Differenz KSA−IRBA")
szenarien = [sz_a, sz_b, sz_c]


# ═══════════════════════════════════════════════════════════════════════════════
# 5. KONSOLENAUSGABE
# ═══════════════════════════════════════════════════════════════════════════════

def trunc(s, n=30):
    return s if len(s) <= n else s[:n-1] + "…"

print("=" * 90)
print("ESSEN BANK – AUFGABE 3.3: OUTPUT FLOOR ANALYSE")
print("=" * 90)

# Tabellenkopf
print(f"\n{'#':>3} {'Name':<32} {'Rat':>6} {'PD%':>5} {'LGD%':>5} {'M':>4} "
      f"{'EAD':>6} {'IRBA-RW%':>9} {'IRBA-RWA':>9} {'KSA-RW%':>8} {'KSA-RWA':>8} {'Diff':>8}")
print("-" * 110)

for p in ergebnisse:
    print(f"{p['id']:>3} {trunc(p['name']):<32} {p['rating']:>6} "
          f"{p['pd_pct']:>5.2f} {p['lgd_pct']:>5.1f} {p['M']:>4.1f} "
          f"{p['ead']:>6.1f} {p['irba_rw']*100:>9.2f} {p['irba_rwa']:>9.3f} "
          f"{p['ksa_rw']*100:>8.0f} {p['ksa_rwa']:>8.3f} {p['differenz']:>8.3f}")

print("-" * 110)
b = baseline
print(f"\n{'PORTFOLIO-SUMME':>50}  EAD={b['ead_sum']:.1f}  "
      f"IRBA-RWA={b['irba_sum']:.2f}  KSA-RWA={b['ksa_sum']:.2f}")
print(f"{'Floor-Basis (72,5% × KSA)':>50}  {b['floor_basis']:.2f} Mio. €")
print(f"{'Floor-RWA = MAX(IRBA, Floor-Basis)':>50}  {b['floor_rwa']:.2f} Mio. €")
print(f"{'Floor bindet?':>50}  {'JA ▲' if b['bindet'] else 'NEIN ✓'}")

print("\n" + "=" * 90)
print("SZENARIEN")
print("=" * 90)

for sz in szenarien:
    print(f"\n{sz['label']}")
    print("  Verkaufte Positionen:")
    for p in sz["verkauft"]:
        print(f"    ID {p['id']:>2}  {trunc(p['name'],28):<28}  "
              f"EAD={p['ead']:5.1f}  IRBA-RWA={p['irba_rwa']:6.3f}  "
              f"KSA-RWA={p['ksa_rwa']:6.3f}  Diff={p['differenz']:+6.3f}")
    v = sz
    delta = v["floor_rwa"] - baseline["floor_rwa"]
    print(f"  Verbleibend: EAD={v['ead_sum']:.1f}  "
          f"IRBA={v['irba_sum']:.2f}  Floor-Basis={v['floor_basis']:.2f}  "
          f"Floor-RWA={v['floor_rwa']:.2f}  bindet={'JA' if v['bindet'] else 'NEIN'}  "
          f"Δ vs Baseline={delta:+.2f}")


# ═══════════════════════════════════════════════════════════════════════════════
# 6. EXCEL-EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def excel_export(ergebnisse, baseline, szenarien, pfad="Essen_Bank_Ergebnisse.xlsx"):

    wb = Workbook()

    # ── Hilfsfunktionen Styling ──────────────────────────────────────────────
    C_H, C_SUB = "2D2B6E", "534AB7"
    C_YEL, C_GRN, C_BLU = "FFFDE7", "E8F5E9", "E3F2FD"
    C_RED, C_TOT = "FAEEDA", "D8D6F8"

    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfill(c):
        return PatternFill("solid", fgColor=c)

    def st(ws, rc, val="", bold=False, color="000000", sz=9,
           bg="FFFFFF", h="center", wrap=False, nf=None, italic=False):
        c = ws[rc]
        c.value     = val
        c.font      = Font(name="Arial", bold=bold, color=color,
                           size=sz, italic=italic)
        c.fill      = hfill(bg)
        c.alignment = Alignment(horizontal=h, vertical="center",
                                wrap_text=wrap)
        c.border    = brd
        if nf:
            c.number_format = nf
        return c

    def titel(ws, span, txt):
        ws.merge_cells(span)
        c = ws[span.split(":")[0]]
        c.value = txt
        c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill  = hfill(C_H)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = brd

    def zeile_h(ws, r, h): ws.row_dimensions[r].height  = h
    def col_w(ws,  c, w): ws.column_dimensions[c].width = w

    # ── SHEET 1: Portfolio ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Portfolio"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A4"

    titel(ws1, "A1:N1",
          "Essen Bank – IRBA & KSA Berechnung  |  Art. 153 CRR  |  50 Unternehmenskredite")
    zeile_h(ws1, 1, 24)

    hdrs = [("A","#",5),("B","Name",32),("C","Branche",14),
            ("D","PD\n(%)",8),("E","LGD\n(%)",7),("F","M\n(J.)",6),
            ("G","EAD\n(Mio.)",9),("H","KMU",5),("I","S\n(Mio.)",8),
            ("J","Rating",9),
            ("K","IRBA-RW\n(%)",11),("L","IRBA-RWA\n(Mio.)",12),
            ("M","KSA-RW\n(%)",10),("N","KSA-RWA\n(Mio.)",12)]
    for col, label, width in hdrs:
        c = ws1[f"{col}3"]
        c.value = label
        c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill  = hfill(C_SUB)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = brd
        col_w(ws1, col, width)
    zeile_h(ws1, 3, 30)

    RATING_FARBEN = {
        "A":       ("E6F1FB","0C447C"),
        "BBB":     ("EEEDFE","3C3489"),
        "BB":      ("FAEEDA","633806"),
        "B":       ("FCEBEB","791F1F"),
        "unrated": ("F1EFE8","444441"),
    }

    DR = 4
    for i, p in enumerate(ergebnisse):
        r   = DR + i
        rbg = "FFFFFF" if i % 2 == 0 else "F8F8FB"

        st(ws1, f"A{r}", p["id"],      bg=rbg, h="center")
        st(ws1, f"B{r}", p["name"],    bg=rbg, h="left", sz=9)
        st(ws1, f"C{r}", p["branche"], bg=rbg, h="left", sz=9)
        st(ws1, f"D{r}", p["pd_pct"]/100,  bg=C_YEL, bold=True,
           color="0000FF", nf="0.00%")
        st(ws1, f"E{r}", p["lgd_pct"]/100, bg=C_YEL, bold=True,
           color="0000FF", nf="0%")
        st(ws1, f"F{r}", p["M"],       bg=C_YEL, bold=True,
           color="0000FF", nf="0.0")
        st(ws1, f"G{r}", p["ead"],     bg=C_YEL, bold=True,
           color="0000FF", nf="#,##0.0")
        st(ws1, f"H{r}", 1 if p["kmu"] else 0, bg=C_YEL,
           bold=True, color="0000FF")
        st(ws1, f"I{r}", p["S_mio"] or "", bg=C_YEL,
           bold=True, color="0000FF", nf="#,##0.0")

        rbg_rt, rfg_rt = RATING_FARBEN.get(p["rating"], ("F1EFE8","444441"))
        c = ws1[f"J{r}"]
        c.value = p["rating"]
        c.font  = Font(name="Arial", bold=True, color=rfg_rt, size=9)
        c.fill  = hfill(rbg_rt)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd

        gbg = "E8F5E9" if i % 2 == 0 else "D4EDD8"
        bbg = "E3F2FD" if i % 2 == 0 else "D0E8F8"
        st(ws1, f"K{r}", p["irba_rw"],  bg=gbg, nf="0.00%")
        st(ws1, f"L{r}", p["irba_rwa"], bg=gbg, nf="#,##0.000")
        st(ws1, f"M{r}", p["ksa_rw"],   bg=bbg, nf="0%")
        st(ws1, f"N{r}", p["ksa_rwa"],  bg=bbg, nf="#,##0.000")
        zeile_h(ws1, r, 15)

    # Summenzeile
    tr = DR + 50
    for col in "ABCDEFGHIJKLMN":
        c = ws1[f"{col}{tr}"]
        c.fill   = hfill(C_TOT)
        c.border = Border(
            top=Side(style="medium", color=C_SUB),
            bottom=Side(style="medium", color=C_SUB),
            left=thin, right=thin)
        c.font      = Font(name="Arial", bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws1[f"B{tr}"].value = "Summe Portfolio"
    ws1[f"B{tr}"].alignment = Alignment(horizontal="left", vertical="center")
    for col, val, nf in [
        ("G", baseline["ead_sum"],  "#,##0.0"),
        ("L", baseline["irba_sum"], "#,##0.000"),
        ("N", baseline["ksa_sum"],  "#,##0.000"),
    ]:
        ws1[f"{col}{tr}"].value          = val
        ws1[f"{col}{tr}"].number_format  = nf
    zeile_h(ws1, tr, 18)

    # Output Floor Zeilen
    for off, label, val, nf, bg, tcol in [
        (1, "Floor-Basis  =  72,5 % × KSA-RWA",
         baseline["floor_basis"], "#,##0.000", C_RED, "854F0B"),
        (2, "Floor-RWA  =  MAX(IRBA-RWA ; Floor-Basis)",
         baseline["floor_rwa"],   "#,##0.000", "FFE0B2", "A32D2D"),
        (3, "Floor bindet?",
         "JA – Floor wirksam ▲" if baseline["bindet"] else "NEIN – IRBA bindend ✓",
         "@", "F1F8E9",
         "A32D2D" if baseline["bindet"] else "3B6D11"),
    ]:
        r2 = tr + off
        ws1.merge_cells(f"A{r2}:J{r2}")
        c = ws1[f"A{r2}"]
        c.value = label
        c.font  = Font(name="Arial", bold=True, size=9, color="3C3289")
        c.fill  = hfill(bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = brd
        vc = ws1[f"K{r2}"]
        vc.value  = val
        vc.number_format = nf
        vc.font   = Font(name="Arial", bold=True, size=10, color=tcol)
        vc.fill   = hfill(bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = brd
        ws1.merge_cells(f"K{r2}:N{r2}")
        zeile_h(ws1, r2, 17)

    # ── SHEET 2: Szenarien ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Szenarien")
    ws2.sheet_view.showGridLines = False

    titel(ws2, "A1:L1",
          "Essen Bank – Szenarioanalyse Output Floor  |  Aufgabe 3.3 a) / b) / c)")
    zeile_h(ws2, 1, 24)

    for col, w in [("A",5),("B",32),("C",9),("D",8),("E",10),
                   ("F",10),("G",10),("H",10),("I",8),("J",10),("K",12),("L",14)]:
        col_w(ws2, col, w)

    SZ_FARBEN = [("FFF3E0","854F0B"), ("FDE8E8","A32D2D"), ("E8F5E9","3B6D11")]
    cur = 3

    for si, sz in enumerate(szenarien):
        hbg, hfg = SZ_FARBEN[si]

        # Abschnittsüberschrift
        ws2.merge_cells(f"A{cur}:L{cur}")
        c = ws2[f"A{cur}"]
        c.value = sz["label"]
        c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill  = hfill(C_SUB)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = brd
        zeile_h(ws2, cur, 18)
        cur += 1

        # Spaltenköpfe
        for col, lbl in [("A","#"),("B","Name"),("C","Rating"),
                         ("D","PD%"),("E","LGD%"),("F","M"),("G","EAD"),
                         ("H","IRBA-RW%"),("I","IRBA-RWA"),
                         ("J","KSA-RW%"),("K","KSA-RWA"),("L","Status")]:
            c = ws2[f"{col}{cur}"]
            c.value = lbl
            c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            c.fill  = hfill("7B75CF")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = brd
        zeile_h(ws2, cur, 18)
        cur += 1

        sold_ids = {p["id"] for p in sz["verkauft"]}
        for p in ergebnisse:
            sold = p["id"] in sold_ids
            rbg  = hbg if sold else ("FAFAFA" if cur % 2 == 0 else "FFFFFF")
            fnt  = hfg if sold else "000000"
            bold = sold

            st(ws2, f"A{cur}", p["id"],        bg=rbg, bold=bold, color=fnt)
            st(ws2, f"B{cur}", p["name"],       bg=rbg, bold=bold, color=fnt, h="left")
            st(ws2, f"C{cur}", p["rating"],     bg=rbg, bold=bold, color=fnt)
            st(ws2, f"D{cur}", p["pd_pct"],     bg=rbg, bold=bold, color=fnt, nf="0.00")
            st(ws2, f"E{cur}", p["lgd_pct"],    bg=rbg, bold=bold, color=fnt, nf="0")
            st(ws2, f"F{cur}", p["M"],          bg=rbg, bold=bold, color=fnt, nf="0.0")
            st(ws2, f"G{cur}", p["ead"],        bg=rbg, bold=bold, color=fnt, nf="#,##0.0")
            st(ws2, f"H{cur}", p["irba_rw"],    bg=rbg, bold=bold, color=fnt, nf="0.00%")
            st(ws2, f"I{cur}", p["irba_rwa"],   bg=rbg, bold=bold, color=fnt, nf="#,##0.000")
            st(ws2, f"J{cur}", p["ksa_rw"],     bg=rbg, bold=bold, color=fnt, nf="0%")
            st(ws2, f"K{cur}", p["ksa_rwa"],    bg=rbg, bold=bold, color=fnt, nf="#,##0.000")
            st(ws2, f"L{cur}", "◀ VERKAUFT" if sold else "",
               bg=rbg, bold=bold, color=hfg if sold else "AAAAAA", sz=8)
            zeile_h(ws2, cur, 14)
            cur += 1

        # Zwischensummen
        v_irba = sum(p["irba_rwa"] for p in sz["verkauft"])
        v_ksa  = sum(p["ksa_rwa"]  for p in sz["verkauft"])
        v_ead  = sum(p["ead"]      for p in sz["verkauft"])

        for lbl, ead_v, irba_v, ksa_v, bg in [
            ("Verkauft (10 Pos.)",    v_ead,         v_irba,       v_ksa,        "FFE0B2"),
            ("Verbleibend (40 Pos.)", sz["ead_sum"], sz["irba_sum"],sz["ksa_sum"],"E3F2FD"),
        ]:
            ws2.merge_cells(f"A{cur}:F{cur}")
            c = ws2[f"A{cur}"]
            c.value = lbl
            c.font  = Font(name="Arial", bold=True, size=9)
            c.fill  = hfill(bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = brd
            for col in "BCDEF":
                ws2[f"{col}{cur}"].fill   = hfill(bg)
                ws2[f"{col}{cur}"].border = brd
            st(ws2, f"G{cur}", ead_v,  bg=bg, bold=True, nf="#,##0.0")
            st(ws2, f"I{cur}", irba_v, bg=bg, bold=True, nf="#,##0.000")
            st(ws2, f"K{cur}", ksa_v,  bg=bg, bold=True, nf="#,##0.000")
            for col in "HJL":
                ws2[f"{col}{cur}"].fill   = hfill(bg)
                ws2[f"{col}{cur}"].border = brd
            zeile_h(ws2, cur, 16)
            cur += 1

        # Floor-Ergebnis
        fl_bg  = C_RED if sz["bindet"] else "EAF3DE"
        fl_col = "A32D2D" if sz["bindet"] else "3B6D11"
        ws2.merge_cells(f"A{cur}:L{cur}")
        c = ws2[f"A{cur}"]
        c.value = (f"Floor-Basis = {sz['floor_basis']:.2f} Mio.  |  "
                   f"Floor-RWA = {sz['floor_rwa']:.2f} Mio.  |  "
                   f"Δ vs. Baseline = {sz['floor_rwa']-baseline['floor_rwa']:+.2f} Mio.  |  "
                   f"Floor bindet: {'JA ▲' if sz['bindet'] else 'NEIN ✓'}")
        c.font  = Font(name="Arial", bold=True, size=9, color=fl_col)
        c.fill  = hfill(fl_bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = brd
        zeile_h(ws2, cur, 17)
        cur += 3

    wb.save(pfad)
    print(f"\n✓  Excel gespeichert: {pfad}")


# ── Export ausführen ──────────────────────────────────────────────────────────
excel_export(ergebnisse, baseline, szenarien,
             pfad="Essen_Bank_Ergebnisse.xlsx")
